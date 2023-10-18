import os, utils
from zipfile import ZipFile
from PyPDF2 import PdfReader
from io import BytesIO
from xlrd import open_workbook
from openpyxl import load_workbook


# проверка txt
def test_txt_file():
    with ZipFile(utils.ZIP_PATH, mode='r') as zf:
        with zf.open('book.txt', 'r') as txt_file:
            f = open(os.path.join(utils.RESOURCES_PATH, 'book.txt'), 'r')
            assert (txt_file.read().decode('utf-8')) == f.read()


# проверка xls
def test_xls_file():
    with ZipFile(utils.ZIP_PATH, mode='r') as zf:
        zip_book = open_workbook(file_contents=zf.read("menu.xls"))
        res_book = open_workbook(os.path.join(utils.RESOURCES_PATH, 'menu.xls'))

        sheets_number_zip = zip_book.nsheets
        sheets_number_res = res_book.nsheets
        assert sheets_number_zip == sheets_number_res

        sheet_zip = zip_book.sheet_by_index(0)
        sheet_res = res_book.sheet_by_index(0)
        assert sheet_zip.cell_value(13, 5) == sheet_res.cell_value(13, 5)
        assert sheet_zip.cell_value(33, 3) == sheet_res.cell_value(33, 3)


#проверка xlsx
def test_xlsx_file():
    with ZipFile(utils.ZIP_PATH, mode='r') as zf:
        zip_workbook = load_workbook(zf.open('orders.xlsx'))
        res_workbook = load_workbook(os.path.join(utils.RESOURCES_PATH, 'orders.xlsx'))
        assert zip_workbook.sheetnames == res_workbook.sheetnames

        zip_xlsx_size = zf.getinfo('orders.xlsx').file_size
        res_xlsx_size = os.path.getsize(os.path.join(utils.RESOURCES_PATH, 'orders.xlsx'))
        assert zip_xlsx_size == res_xlsx_size

        sheet_zip_for_xlsx = zip_workbook.active
        sheet_res_for_xlsx = res_workbook.active
        assert sheet_zip_for_xlsx.cell(8, 5).value == sheet_res_for_xlsx.cell(8, 5).value


# проверка pdf
def test_pdf_file():
    with ZipFile(utils.ZIP_PATH, mode='r') as zf:
        zip_pdf = PdfReader(BytesIO(zf.read('SQL.pdf')))
        resources_pdf = PdfReader(os.path.join(utils.RESOURCES_PATH, 'SQL.pdf'))
        assert len(zip_pdf.pages) == len(resources_pdf.pages)
        assert zip_pdf.pages[3].extract_text() == resources_pdf.pages[3].extract_text()

